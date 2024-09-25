import json
import codecs
import os
import subprocess

try:
    import openpyxl
except ImportError:
    subprocess.call(['pip', 'install', 'openpyxl'])

from openpyxl import load_workbook

def ExcelToJson():
    dirpath = os.listdir(os.path.abspath(os.curdir))
    result = {}
    tab = "Excel列表 " + '\n'
    num = 0
    for file_path in dirpath:
        if file_path.endswith(".xlsx"):
            wb = load_workbook(file_path)
            if wb is not None:
                # 获取所有工作表的名字
                worksheets = wb.sheetnames
                for sheet_name in worksheets:
                    # 记录列表
                    num += 1
                    tab += f'{num},{sheet_name}\n'
                    table = wb[sheet_name]
                    # 初始化记录Json的部分
                    result["num"] = num
                    result[str(sheet_name)] = []
                    # 获取第一行标题并跳过
                    row_0 = [cell.value for cell in next(table.rows)]
                    # 遍历每一行数据
                    for row in table.iter_rows(min_row=2):  # 从第二行开始遍历
                        tmp = {}
                        for idx, cell in enumerate(row):
                            title_cn = row_0[idx]
                            tmp[title_cn] = cell.value
                        result[str(sheet_name)].append(tmp)
    json_data = json.dumps(result, indent=4, ensure_ascii=False)
    saveFile(os.getcwd(), "Excel列表", tab, ".txt")
    saveFile(os.getcwd(), "JsonData", json_data, ".json")
    print(json_data)
    return json_data

# 获取excel数据源
def get_data(file_path):
    """获取excel数据源"""
    try:
        data = xlrd.open_workbook(file_path)
        return data
    except Exception:
        print('excel表格读取失败')
        return None


def saveFile(file_path, file_name, data, pre):
    output = codecs.open(file_path + "/" + file_name + pre, 'w', "utf-8")
    output.write(data)
    output.close()


if __name__ == '__main__':
    jsdata = ExcelToJson()
